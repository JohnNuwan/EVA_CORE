#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IO Hardware Ingester - Outil d'ingestion de listes d'E/S industrielles.
Génère des fichiers de variables pour Siemens (SCL), Schneider (XMY), ou PLCopen XML.
"""

import sys
import os
import csv
import argparse
from pathlib import Path
import xml.etree.ElementTree as ET
from xml.dom import minidom
from datetime import datetime

# Essayer d'importer openpyxl pour le support d'Excel
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

def prettify_xml(elem: ET.Element) -> str:
    """Retourne une chaîne XML formatée de façon lisible."""
    rough_string = ET.tostring(elem, 'utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="  ", encoding="utf-8").decode("utf-8")

class IOIngester:
    def __init__(self, input_path: Path):
        self.input_path = input_path
        self.io_list = []

    def parse(self) -> bool:
        """Détermine le format (CSV ou Excel) et extrait les données d'E/S."""
        if not self.input_path.exists():
            print(f"Erreur : Le fichier d'entrée '{self.input_path}' n'existe pas.", file=sys.stderr)
            return False

        suffix = self.input_path.suffix.lower()
        if suffix == ".csv":
            return self._parse_csv()
        elif suffix in (".xlsx", ".xls"):
            if not HAS_OPENPYXL:
                print("Avertissement : openpyxl n'est pas installé. Tentative d'installation automatique...", file=sys.stderr)
                import subprocess
                try:
                    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                    global openpyxl, HAS_OPENPYXL
                    import openpyxl
                    HAS_OPENPYXL = True
                except Exception as e:
                    print(f"Erreur : Impossible d'installer openpyxl ({e}). Convertissez d'abord votre fichier en CSV.", file=sys.stderr)
                    return False
            return self._parse_excel()
        else:
            print(f"Erreur : Format de fichier non supporté '{suffix}'. Utilisez .csv ou .xlsx.", file=sys.stderr)
            return False

    def _parse_csv(self) -> bool:
        """Parse un fichier CSV contenant des colonnes d'E/S."""
        try:
            # Essayer de deviner le délimiteur (souvent ';' en France pour Excel)
            with open(self.input_path, "r", encoding="utf-8", errors="ignore") as f:
                sample = f.read(2048)
                f.seek(0)
                
                # Détecter le délimiteur
                delimiter = ";"
                if sample:
                    if "," in sample and sample.count(",") > sample.count(";"):
                        delimiter = ","
                    elif "\t" in sample:
                        delimiter = "\t"
                        
                reader = csv.DictReader(f, delimiter=delimiter)
                # Normaliser les noms de champs
                fieldnames = [fn.strip().lower() for fn in reader.fieldnames] if reader.fieldnames else []
                
                # Mapper les colonnes clés (tag, adresse, type, description)
                col_mapping = self._detect_columns(fieldnames)
                if not col_mapping["tag"]:
                    print("Erreur : Impossible de trouver une colonne pour le 'Tag' / Mnémonique dans le CSV.", file=sys.stderr)
                    print(f"Colonnes trouvées : {reader.fieldnames}", file=sys.stderr)
                    return False
                
                # Lire chaque ligne
                for row in reader:
                    # Nettoyer la ligne
                    clean_row = {k.strip().lower(): v.strip() for k, v in row.items() if k is not None}
                    
                    tag = clean_row.get(col_mapping["tag"])
                    if not tag:
                        continue # ignorer les lignes sans tag
                        
                    addr = clean_row.get(col_mapping["addr"], "")
                    sig_type = clean_row.get(col_mapping["type"], "")
                    desc = clean_row.get(col_mapping["desc"], "")
                    
                    self.io_list.append({
                        "tag": tag,
                        "address": addr,
                        "type": self._normalize_type(sig_type, addr),
                        "description": desc
                    })
            return True
        except Exception as e:
            print(f"Erreur lors de la lecture du fichier CSV : {e}", file=sys.stderr)
            return False

    def _parse_excel(self) -> bool:
        """Parse un fichier Excel .xlsx."""
        try:
            wb = openpyxl.load_workbook(self.input_path, data_only=True)
            sheet = wb.active
            
            # Découvrir l'en-tête (première ligne non vide)
            header_row = 1
            headers = []
            for r in range(1, 10):
                row_vals = [str(sheet.cell(r, c).value or "").strip() for c in range(1, sheet.max_column + 1)]
                # Si au moins deux cellules contiennent des mots-clés de listes d'E/S
                keywords = ["tag", "address", "adresse", "type", "description", "mnemonique"]
                matches = sum(1 for v in row_vals if any(kw in v.lower() for kw in keywords))
                if matches >= 2:
                    header_row = r
                    headers = [v.lower() for v in row_vals]
                    break
            
            if not headers:
                # Fallback par défaut sur la première ligne
                headers = [str(sheet.cell(1, c).value or "").strip().lower() for c in range(1, sheet.max_column + 1)]
                header_row = 1
                
            col_mapping = self._detect_columns(headers)
            if not col_mapping["tag"]:
                print("Erreur : Impossible de trouver une colonne pour le 'Tag' / Mnémonique dans l'Excel.", file=sys.stderr)
                print(f"Colonnes d'en-tête : {headers}", file=sys.stderr)
                return False
                
            # Extraire les indices (1-indexed pour openpyxl)
            tag_idx = headers.index(col_mapping["tag"]) + 1
            addr_idx = headers.index(col_mapping["addr"]) + 1 if col_mapping["addr"] in headers else None
            type_idx = headers.index(col_mapping["type"]) + 1 if col_mapping["type"] in headers else None
            desc_idx = headers.index(col_mapping["desc"]) + 1 if col_mapping["desc"] in headers else None
            
            for r in range(header_row + 1, sheet.max_row + 1):
                tag = str(sheet.cell(r, tag_idx).value or "").strip()
                if not tag or tag.lower() in ("none", "null", ""):
                    continue
                    
                addr = str(sheet.cell(r, addr_idx).value or "").strip() if addr_idx else ""
                sig_type = str(sheet.cell(r, type_idx).value or "").strip() if type_idx else ""
                desc = str(sheet.cell(r, desc_idx).value or "").strip() if desc_idx else ""
                
                self.io_list.append({
                    "tag": tag,
                    "address": addr,
                    "type": self._normalize_type(sig_type, addr),
                    "description": desc
                })
            return True
        except Exception as e:
            print(f"Erreur lors de la lecture du fichier Excel : {e}", file=sys.stderr)
            return False

    def _detect_columns(self, headers: list) -> dict:
        """Détecte automatiquement l'association des en-têtes de colonnes."""
        mapping = {"tag": None, "addr": None, "type": None, "desc": None}
        
        # Mots-clés pour chaque colonne recherchée
        keywords = {
            "tag": ["tag", "mnemonique", "nom", "name", "symbole", "identifier"],
            "addr": ["address", "adresse", "addr", "physical", "physique", "loc", "location"],
            "type": ["type", "signal", "dataType", "datatype", "format"],
            "desc": ["desc", "description", "fonction", "comment", "commentaire", "libelle"]
        }
        
        for key, kw_list in keywords.items():
            for h in headers:
                if any(kw in h for kw in kw_list):
                    mapping[key] = h
                    break
                    
        # Fallback de secours si non détecté
        if not mapping["tag"] and headers:
            mapping["tag"] = headers[0] # première colonne par défaut pour le Tag
        if not mapping["addr"] and len(headers) > 1:
            mapping["addr"] = headers[1]
        if not mapping["type"] and len(headers) > 2:
            mapping["type"] = headers[2]
        if not mapping["desc"] and len(headers) > 3:
            mapping["desc"] = headers[3]
            
        return mapping

    def _normalize_type(self, sig_type: str, addr: str) -> str:
        """Normalise le type de données automate à partir du type ou de l'adresse."""
        t_upper = sig_type.upper()
        if "DI" in t_upper or "DO" in t_upper or "BOOL" in t_upper:
            return "BOOL"
        if "AI" in t_upper or "AO" in t_upper or "INT" in t_upper or "WORD" in t_upper:
            return "INT"
            
        # Détection par l'adresse Siemens/Rockwell
        addr_upper = addr.upper()
        if "%I" in addr_upper or "%Q" in addr_upper:
            if "W" in addr_upper or "D" in addr_upper:
                return "INT"
            return "BOOL"
            
        return "BOOL" # Par défaut

    def generate_scl(self) -> str:
        """Génère un bloc de déclaration Siemens SCL (Data Block)."""
        lines = []
        lines.append('DATA_BLOCK "IO_DB"')
        lines.append('{ S7_Optimized_Access := \'TRUE\' }')
        lines.append('VERSION : 0.1')
        lines.append('VAR')
        
        for item in self.io_list:
            comment_part = f" // {item['description']}" if item['description'] else ""
            # Siemens accepte des adresses physiques dans des déclarations de variables globales (ex: %I0.0)
            # uniquement sous forme de tags PLC (non DB). Si c'est un DB, c'est juste de la mémoire.
            # Pour faire un fichier propre, on met l'adresse en commentaire si besoin, ou on fait une déclaration standard.
            addr_comment = f" [Adresse: {item['address']}]" if item['address'] else ""
            lines.append(f"   {item['tag']} : {item['type']};{comment_part}{addr_comment}")
            
        lines.append('END_VAR')
        lines.append('BEGIN')
        lines.append('END_DATA_BLOCK')
        return "\n".join(lines)

    def generate_xmy(self) -> str:
        """Génère un fichier d'importation de variables Schneider Unity/Control Expert (.xmy)."""
        root = ET.Element("VariablesExchangeFile")
        
        file_header = ET.SubElement(root, "FileHeader", 
                                    company="Actemium", 
                                    product="Helios IO Ingester", 
                                    version="1.0")
        
        variables_db = ET.SubElement(root, "VariablesDB")
        
        for item in self.io_list:
            # Schneider utilise EBOOL pour le booléen physique avec gestion d'historique/fronts, et INT
            schneider_type = "EBOOL" if item['type'] == "BOOL" else "INT"
            
            var_attrs = {"name": item['tag'], "type": schneider_type}
            if item['address']:
                # Traduire les adresses Siemens type %I1.0 en Schneider %I1.0 ou %MW
                var_attrs["address"] = item['address']
                
            var_el = ET.SubElement(variables_db, "Variable", **var_attrs)
            
            if item['description']:
                comment_el = ET.SubElement(var_el, "Comment")
                comment_el.text = item['description']
                
        return prettify_xml(root)

    def generate_plcopen_xml(self) -> str:
        """Génère un fichier au standard universel PLCopen XML (pour CODESYS/TwinCAT)."""
        project = ET.Element("project", xmlns="http://www.plcopen.org/xml/tc6_0201")
        
        # En-têtes
        file_header = ET.SubElement(project, "fileHeader", 
                                    companyName="Actemium", 
                                    productName="Helios IO Ingester", 
                                    productVersion="1.0", 
                                    creationDateTime=datetime.now().strftime("%Y-%m-%dT%H:%M:%S"))
        
        content_header = ET.SubElement(project, "contentHeader", name="IO_Variables")
        
        types = ET.SubElement(project, "types")
        global_vars = ET.SubElement(types, "globalVars", name="IO_GVL")
        
        for item in self.io_list:
            var_el = ET.SubElement(global_vars, "variable", name=item['tag'])
            
            if item['address']:
                var_el.set("address", item['address'])
                
            type_el = ET.SubElement(var_el, "type")
            ET.SubElement(type_el, item['type'])
            
            if item['description']:
                doc_el = ET.SubElement(var_el, "documentation")
                html_el = ET.SubElement(doc_el, "{http://www.w3.org/1999/xhtml}html")
                p_el = ET.SubElement(html_el, "p")
                p_el.text = item['description']
                
        return prettify_xml(project)

    def export(self, format_name: str, output_path: Path) -> bool:
        """Exporte les variables au format désiré dans le fichier spécifié."""
        if not self.io_list:
            print("Avertissement : Aucune variable d'E/S à exporter.", file=sys.stderr)
            
        format_name = format_name.lower()
        if format_name == "scl":
            content = self.generate_scl()
        elif format_name == "xmy":
            content = self.generate_xmy()
        elif format_name == "plcopen":
            content = self.generate_plcopen_xml()
        else:
            print(f"Erreur : Format d'exportation non supporté '{format_name}'.", file=sys.stderr)
            return False
            
        try:
            # Créer les dossiers parents si nécessaire
            output_path.parent.mkdir(parents=True, exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(content)
            print(f"Succès : Fichier exporté avec succès ({len(self.io_list)} variables) vers '{output_path}'.")
            return True
        except Exception as e:
            print(f"Erreur lors de l'écriture du fichier de sortie : {e}", file=sys.stderr)
            return False

def main():
    parser = argparse.ArgumentParser(description="Ingère des listes d'E/S industrielles et génère des variables PLC.")
    parser.add_argument("input", help="Fichier d'entrée CSV ou Excel (.csv, .xlsx)")
    parser.add_argument("--output", "-o", required=True, help="Chemin du fichier de sortie généré")
    parser.add_argument("--format", "-f", required=True, choices=["scl", "xmy", "plcopen"], help="Format de sortie des variables")
    
    args = parser.parse_args()
    
    input_path = Path(args.input)
    output_path = Path(args.output)
    
    ingester = IOIngester(input_path)
    if not ingester.parse():
        sys.exit(1)
        
    if not ingester.export(args.format, output_path):
        sys.exit(1)

if __name__ == "__main__":
    main()
